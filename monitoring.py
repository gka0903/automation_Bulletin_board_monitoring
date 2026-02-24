import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from copy import copy
from datetime import datetime
import os
import time
import re

from config import AGENCY_CONFIG

# 엑셀 파일 경로
EXCEL_FILE = r"C:\Users\default.DESKTOP-JNPQLP7\Desktop\업무\게시판 모니터링 인수인계\최신화 ★2026_주요기관 게시판 및 디전_사전협의 모니터링_최신.xlsx"


def read_excel_data(file_path, sheet_name):
    # 엑셀에서 마지막에 수집된 사업명을 가져옵니다.
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df['사업명'].iloc[-1]
    except Exception as e:
        print(f"❌ 엑셀 읽기 오류 (Excel read error): {e}")
        return None


def start_monitoring(driver, target_name, config):
    # 설정값을 바탕으로 게시판을 탐색합니다.
    driver.get(config["url"])
    new_post_links = []
    sel = config["selectors"]

    while True:
        time.sleep(3)

        try:
            li_elements = driver.find_elements(By.CSS_SELECTOR, sel["list_item"])
            stop_searching = False

            for li in li_elements:
                try:
                    title_text = li.find_element(By.CSS_SELECTOR, sel["title"]).text.strip()

                    if target_name in title_text:
                        print(f"🏁 기준점 발견 (Found reference point) [{title_text}]. 탐색 중단 (Stopping search).")
                        stop_searching = True
                        break

                    a_tag = li.find_element(By.CSS_SELECTOR, sel["link_tag"])
                    link_val = a_tag.get_attribute(config["link_type"])

                    if not any(item['title'] == title_text for item in new_post_links):
                        print(f"🆕 새 공고 수집 (Collected new post): {title_text}")
                        new_post_links.append({
                            'title': title_text,
                            'link_val': link_val
                        })
                except:
                    continue

            if stop_searching:
                break

            # 다음 페이지 이동
            print("👉 현재 페이지에 없음. 다음 페이지로 이동... (Not on current page. Moving to next page...)")
            current_page_num = int(driver.find_element(By.CSS_SELECTOR, sel["active_page"]).text)
            next_page_num = current_page_num + 1

            try:
                driver.find_element(By.LINK_TEXT, str(next_page_num)).click()
            except:
                driver.find_element(By.CSS_SELECTOR, sel["next_page"]).click()

        except Exception as e:
            print(f"⚠️ 검색 중 오류 발생 (Error during search): {e}")
            break

    return new_post_links


def extract_detail_info(driver, config):
    sel = config["selectors"]
    try:
        time.sleep(2)

        # 선택자가 XPath인지 CSS 선택자인지 자동 판별하여 요소를 찾는 헬퍼 함수
        def get_elem(selector_str):
            if not selector_str: return None
            by_type = By.XPATH if selector_str.startswith("//") else By.CSS_SELECTOR
            try:
                return driver.find_element(by_type, selector_str)
            except:
                return None

        def get_elems(selector_str):
            if not selector_str: return []
            by_type = By.XPATH if selector_str.startswith("//") else By.CSS_SELECTOR
            try:
                return driver.find_elements(by_type, selector_str)
            except:
                return []

        # 1) 제목 및 등록일 추출
        title_elem = get_elem(sel["detail_title"])
        raw_title = " ".join(title_elem.text.split()).strip() if title_elem else ""
        title = re.sub(r'\[.*?\]', '', raw_title).strip()

        date_elem = get_elem(sel["detail_date"])
        if date_elem:
            raw_date_text = date_elem.text.strip()
            try:
                # pd.to_datetime을 사용하면 "2026. 2. 20" 또는 "Feb 23, 2026" 모두 자동으로 파싱합니다.
                parsed_date = pd.to_datetime(raw_date_text)
                reg_date = parsed_date.strftime("%Y-%m-%d")  # "YYYY-MM-DD" 형태로 통일
            except Exception:
                # 자동 변환 실패 시 기본 텍스트 공백 제거 처리
                reg_date = raw_date_text.replace(" ", "").replace(".", "-")
        else:
            reg_date = ""

        # 2) 담당자 및 부서 추출 분기 처리
        if "detail_writer" in sel:

            # NIA 방식: 하나의 선택자로 가져와 인덱스로 분리
            writer_elements = get_elems(sel["detail_writer"])
            manager = writer_elements[0].text.strip() if len(writer_elements) > 0 else ""
            team = writer_elements[1].text.strip() if len(writer_elements) > 1 else ""
        else:

            # KISA 방식: 부서와 담당자가 별도의 선택자로 지정된 경우
            team_elem = get_elem(sel.get("detail_team", ""))
            team = team_elem.text.strip() if team_elem else ""

            manager_elem = get_elem(sel.get("detail_manager", ""))
            manager = manager_elem.text.strip() if manager_elem else ""

        return {"title": title, "reg_date": reg_date, "manager": manager, "team": team}

    except Exception as e:
        print(f"❌ 상세 정보 추출 실패 (Failed to extract detail info): {e}")
        return None


def update_excel_results(file_path, sheet_name, data_list):
    # 엑셀 업데이트 및 서식 복사 로직입니다.
    if not data_list:
        return

    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    today = datetime.now().date()

    actual_last_row = ws.max_row
    while actual_last_row > 1 and not ws.cell(row=actual_last_row, column=6).value:
        actual_last_row -= 1

    print(f"📊 현재 실제 데이터 마지막 행 (Current actual last row): {actual_last_row}")

    for data in reversed(data_list):
        new_row = actual_last_row + 1

        ws.cell(row=new_row, column=1).value = ws.cell(row=actual_last_row, column=1).value
        ws.cell(row=new_row, column=2).value = today
        ws.cell(row=new_row, column=2).number_format = 'yyyy-mm-dd'
        ws.cell(row=new_row, column=3).value = ws.cell(row=actual_last_row, column=3).value
        ws.cell(row=new_row, column=4).value = data['team']
        ws.cell(row=new_row, column=5).value = data['manager']
        ws.cell(row=new_row, column=6).value = data['title']
        # 공고 등록일 (Column 8) - 문자열을 날짜 객체로 변환
        try:
            # pd.to_datetime을 사용하여 문자열을 날짜 객체로 변환 후 date()만 추출
            reg_date_obj = pd.to_datetime(data['reg_date']).date()
            ws.cell(row=new_row, column=8).value = reg_date_obj
        except:
            # 변환 실패 시 기존 텍스트 그대로 입력
            ws.cell(row=new_row, column=8).value = data['reg_date']

        # 날짜 서식 적용 (Formatting)
        ws.cell(row=new_row, column=8).number_format = 'yyyy-mm-dd'

        for col in range(1, 19):
            source_cell = ws.cell(row=actual_last_row, column=col)
            new_cell = ws.cell(row=new_row, column=col)
            if source_cell.has_style:
                new_cell.font = copy(source_cell.font)
                new_cell.border = copy(source_cell.border)
                new_cell.fill = copy(source_cell.fill)
                new_cell.number_format = copy(source_cell.number_format)
                new_cell.alignment = copy(source_cell.alignment)

        actual_last_row += 1

    wb.save(file_path)
    print(f"💾 엑셀 저장 완료 (Excel save completed): {len(data_list)}건 추가됨. ({len(data_list)} items added.)")


def main():
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.maximize_window()

    # 설정된 모든 기관을 순회하며 모니터링을 진행합니다.
    for agency_name, config in AGENCY_CONFIG.items():
        print(f"\n========================================")
        print(f"🏢 [{agency_name}] 모니터링 시작 (Starting monitoring)")
        print(f"========================================")

        target_name = read_excel_data(EXCEL_FILE, config["sheet_name"])
        if not target_name:
            print(f"⚠️ {agency_name}의 기준점을 찾을 수 없습니다. (Could not find reference point for {agency_name}.)")
            continue

        print(f"🎯 찾고 있는 사업명 (Target project name): {target_name}")

        links = start_monitoring(driver, target_name, config)
        final_data = []

        if links:
            main_window = driver.current_window_handle
            print(f"🔎 총 {len(links)}개의 신규 공고를 새 탭으로 열어 분석합니다. (Analyzing a total of {len(links)} new posts in new tabs.)")

            for i, item in enumerate(links, 1):
                print(f"🚀 [{i}/{len(links)}] 상세 페이지 여는 중 (Opening detail page): {item['title'][:20]}...")
                try:
                    driver.execute_script("window.open('');")
                    driver.switch_to.window(driver.window_handles[-1])
                    driver.get(config["url"])
                    time.sleep(2)

                    # 링크 타입에 따라 페이지 이동 방식을 다르게 처리합니다.
                    if config["link_type"] == "onclick":
                        driver.execute_script(item['link_val'])
                    else:
                        driver.get(item['link_val'])

                    details = extract_detail_info(driver, config)

                    if details:
                        final_data.append(details)
                        print(f"   ✅ 완료 (Completed): {details['title']}")

                    driver.switch_to.window(main_window)
                    time.sleep(1)
                except Exception as e:
                    print(f"   ❌ 오류 발생 (Error occurred): {e}")
                    driver.switch_to.window(main_window)

            update_excel_results(EXCEL_FILE, config["sheet_name"], final_data)
            print(f"\n✨ [{agency_name}] 총 {len(final_data)}건 수집 완료. (Collected a total of {len(final_data)} items.)")

        else:
            print(f"🤷‍♂️ [{agency_name}] 새로운 공고가 없습니다. (No new posts.)")

    print("📢 모든 기관의 모니터링이 종료되었습니다. (Monitoring for all agencies has ended.)")
    os.startfile(EXCEL_FILE)


if __name__ == "__main__":
    main()